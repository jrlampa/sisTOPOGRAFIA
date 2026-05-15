"""
Script detalhado para examinar o Excel e extrair método de cálculo e dados de postes.
"""

import sys
import io
# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
from pathlib import Path
import pandas as pd

excel_path = Path(r"C:\Users\jonat\Downloads\CALC TRAÇÃO\GD - IMPACT CONSULTORIA LTDA - UFV ESTIMA__PARTE_3_POSTE_22.xlsm")

try:
    wb = load_workbook(excel_path)
    
    print("=" * 100)
    print("ANÁLISE DETALHADA: Procurando método de cálculo e dados de postes")
    print("=" * 100)
    
    # Ler a aba "Ponto (1)" que parece ser dados de cálculo
    print("\n\n📋 ABA: 'Ponto (1)' - DADOS ESTRUTURAIS E CÁLCULOS")
    print("─" * 100)
    
    ws = wb['Ponto (1)']
    print(f"Dimensões: {ws.max_row} linhas × {ws.max_column} colunas\n")
    
    # Ler sem merged cells
    for row_idx in range(1, min(60, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(15, ws.max_column + 1)):
            try:
                cell = ws.cell(row_idx, col_idx)
                # Pular merged cells
                if hasattr(cell, 'value'):
                    val = cell.value
                else:
                    val = None
                
                if val is not None:
                    if isinstance(val, (int, float)):
                        row_data.append(f"{val}")
                    else:
                        row_data.append(str(val)[:40])
            except:
                pass
        
        if row_data:  # Só mostrar linhas com dados
            print(f"Linha {row_idx:3d}: {' | '.join(row_data)}")
    
    # Procurar por palavras-chave relacionadas a postes e tração
    print("\n\n" + "=" * 100)
    print("PROCURANDO TERMOS ESPECÍFICOS")
    print("=" * 100)
    
    keywords = {
        'poste': ['poste', 'pole', 'altura', 'height', 'esforço', 'effort', 'nominal', '8.5', '10.5', '12.5', '14'],
        'tração': ['tração', 'traction', 'tensão', 'tension', 'força', 'force', 'newton', 'n/mm2', 'mpa'],
        'método': ['fórmula', 'formula', 'método', 'method', 'cálculo', 'calculation', '='],
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row_idx, col_idx)
                    val = cell.value
                    
                    if val and isinstance(val, str):
                        val_lower = val.lower()
                        
                        for category, terms in keywords.items():
                            if any(term in val_lower for term in terms):
                                # Mostrar contexto
                                print(f"\n📌 {category.upper()}: {sheet_name} (L{row_idx}:C{col_idx})")
                                print(f"   Valor: {val[:80]}")
                                
                                # Contexto: próximas colunas
                                context_vals = []
                                for cc in range(col_idx, min(col_idx + 4, ws.max_column + 1)):
                                    try:
                                        context_vals.append(str(ws.cell(row_idx, cc).value)[:30])
                                    except:
                                        pass
                                if context_vals:
                                    print(f"   Contexto linha: {' | '.join(context_vals)}")
                                
                                # Mostrar células abaixo (para fórmulas)
                                if category == 'método':
                                    for rr in range(row_idx, min(row_idx + 6, ws.max_row + 1)):
                                        below_val = ws.cell(rr, col_idx).value
                                        if below_val:
                                            print(f"     L{rr}: {below_val}")
                                
                                break
                except:
                    pass
    
    # Extrair dados de "Ponto (1)" de forma estruturada
    print("\n\n" + "=" * 100)
    print("EXTRAINDO DADOS ESTRUTURADOS DE 'Ponto (1)'")
    print("=" * 100)
    
    ws = wb['Ponto (1)']
    
    # Usar pandas para ler a aba
    df = pd.read_excel(excel_path, sheet_name='Ponto (1)', header=None)
    
    # Procurar por células com valores numéricos que possam ser parâmetros
    print("\nCélulas com valores numéricos (potenciais parâmetros de projeto):")
    for idx, row in df.iterrows():
        for col_idx, val in enumerate(row):
            if isinstance(val, (int, float)) and not pd.isna(val):
                # Verificar se há label na coluna ao lado
                label = None
                if col_idx > 0:
                    label = df.iloc[idx, col_idx - 1]
                
                print(f"  {label}: {val} (L{idx + 1}:C{col_idx + 1})")
    
    print("\n✅ Análise concluída")
    
except Exception as e:
    print(f"❌ Erro: {e}")
    import traceback
    traceback.print_exc()
