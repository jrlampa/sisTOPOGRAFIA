from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

# ── Canonical workbook resolution ────────────────────────────────────────────
# Priority order:
#   1. CQT_WORKBOOK_PATH env var
#   2. First *.xlsx found in Light_estudo/
#   3. Legacy fallback: root-level BECO DO MATA 7 workbook
_LEGACY_WORKBOOK = "CQTsimplificado_BECO DO MATA 7 - PARIDADE_FINAL.xlsx"
_LIGHT_ESTUDO_DIR = Path("Light_estudo")


def _resolve_default_workbook() -> str:
    env_path = os.environ.get("CQT_WORKBOOK_PATH", "").strip()
    if env_path:
        return env_path
    if _LIGHT_ESTUDO_DIR.is_dir():
        candidates = sorted(_LIGHT_ESTUDO_DIR.glob("*.xlsx"))
        if candidates:
            return str(candidates[0])
    return _LEGACY_WORKBOOK


DEFAULT_WORKBOOK_PATH = _resolve_default_workbook()
DEFAULT_OUTPUT_PATH = Path("docs") / "CQT_WORKBOOK_AUDIT.json"
COL_REF_RE = re.compile(r"\$?([A-Z]{1,3})\$?\d+")


PARITY_CELL_MAP: dict[str, dict[str, list[str]]] = {
    "atual": {
        "RAMAL": ["AA30"],
        "GERAL": ["P31", "P32"],
        "DB": ["K6", "K7", "K8", "K10"],
    },
    "proj1": {
        "GERAL PROJ": ["P31", "P32"],
    },
    "proj2": {
        "GERAL PROJ2": ["P31", "P32"],
    },
}


def _extract_parity_cells(wb_v: Any) -> dict[str, dict[str, Any]]:
    parity_cells: dict[str, dict[str, Any]] = {}

    for scenario, sheet_cells in PARITY_CELL_MAP.items():
        scenario_values: dict[str, Any] = {}
        for sheet_name, cells in sheet_cells.items():
            if sheet_name not in wb_v.sheetnames:
                continue

            ws = wb_v[sheet_name]
            for cell_ref in cells:
                key = f"{sheet_name}!{cell_ref}"
                value = ws[cell_ref].value
                if isinstance(value, (int, float)):
                    scenario_values[key] = float(value)
                elif isinstance(value, str) and value.startswith("#"):
                    scenario_values[key] = {"error": value}

        parity_cells[scenario] = scenario_values

    return parity_cells


def _build_recommendations(summary: dict[str, Any]) -> list[str]:
    recommendations: list[str] = []

    if summary["formulaRefLiteralCount"] > 0:
        recommendations.append(
            "Fix formulas containing literal #REF! tokens "
            "before relying on parity baselines."
        )

    if summary["errorByType"].get("#REF!", 0) > 0:
        recommendations.append(
            "Map and repair broken references (#REF!) "
            "at source, then refresh calculated values."
        )

    if summary["externalReferenceFormulaCount"] > 0:
        recommendations.append(
            "Replace external workbook links with local/static "
            "inputs for deterministic audits."
        )

    if summary["visibleFormulasReferencingHiddenCols"] > 0:
        recommendations.append(
            "Include hidden columns in extraction/audit pipelines; "
            "visible CQT outputs depend on them."
        )

    if summary["errorCellCount"] > 0:
        recommendations.append(
            "Track parity state as value + error class "
            "(e.g., #VALUE!, #DIV/0!, #REF!) for robustness."
        )

    if not recommendations:
        recommendations.append(
            "Workbook has no major structural risks for parity workflow."
        )

    return recommendations


def _build_risk(summary: dict[str, Any]) -> dict[str, Any]:
    score = 0
    critical_flags: list[str] = []
    warning_flags: list[str] = []

    if summary["formulaRefLiteralCount"] > 0:
        score += 35
        critical_flags.append("formula_ref_literal")

    if summary["errorByType"].get("#REF!", 0) > 0:
        score += 25
        critical_flags.append("cached_ref_errors")

    if summary["externalReferenceFormulaCount"] > 0:
        score += 15
        warning_flags.append("external_links")

    if summary["hiddenColCount"] > 0 and summary["formulaCellCount"] > 0:
        hidden_dependency_ratio = summary["visibleFormulasReferencingHiddenCols"] / max(
            1, summary["visibleFormulaCells"]
        )
        if hidden_dependency_ratio >= 0.3:
            score += 20
            warning_flags.append("hidden_dependency_high")
        elif hidden_dependency_ratio >= 0.1:
            score += 10
            warning_flags.append("hidden_dependency_medium")

    if summary["errorCellCount"] >= 200:
        score += 20
        warning_flags.append("many_cached_errors")
    elif summary["errorCellCount"] >= 50:
        score += 10
        warning_flags.append("some_cached_errors")

    level = "low"
    if score >= 60:
        level = "high"
    elif score >= 30:
        level = "medium"

    return {
        "level": level,
        "score": score,
        "criticalFlags": critical_flags,
        "warningFlags": warning_flags,
        "checkFailed": len(critical_flags) > 0,
    }


def analyze_workbook(workbook_path: Path) -> dict[str, Any]:
    wb_f = load_workbook(
        workbook_path,
        data_only=False,
        keep_vba=True,
        read_only=False,
    )
    wb_v = load_workbook(
        workbook_path,
        data_only=True,
        keep_vba=True,
        read_only=False,
    )

    error_by_type: Counter[str] = Counter()
    sheet_stats: list[dict[str, Any]] = []

    formula_cell_count = 0
    visible_formula_cells = 0
    hidden_row_count = 0
    hidden_col_count = 0
    error_cell_count = 0
    formula_ref_literal_count = 0
    external_reference_formula_count = 0
    visible_formulas_referencing_hidden_cols = 0

    formula_ref_literal_sample: list[dict[str, str]] = []
    error_sample: list[dict[str, str]] = []
    external_sample: list[dict[str, str]] = []

    for ws in wb_f.worksheets:
        ws_v = wb_v[ws.title]

        sheet_hidden_rows = sum(1 for _, dim in ws.row_dimensions.items() if dim.hidden)
        sheet_hidden_cols = [
            col for col, dim in ws.column_dimensions.items() if dim.hidden
        ]

        hidden_row_count += sheet_hidden_rows
        hidden_col_count += len(sheet_hidden_cols)

        sheet_formula_cells = 0
        sheet_visible_formula_cells = 0
        sheet_error_cells = 0

        hidden_cols_set = set(sheet_hidden_cols)

        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                formula = cell.value
                cached = ws_v[cell.coordinate].value
                if not isinstance(cell.column, int):
                    continue

                col_letter = get_column_letter(cell.column)
                is_hidden_col = col_letter in hidden_cols_set

                if isinstance(formula, str) and formula.startswith("="):
                    sheet_formula_cells += 1
                    formula_cell_count += 1

                    if not is_hidden_col:
                        sheet_visible_formula_cells += 1
                        visible_formula_cells += 1

                    upper_formula = formula.upper()

                    if "#REF!" in upper_formula:
                        formula_ref_literal_count += 1
                        if len(formula_ref_literal_sample) < 30:
                            formula_ref_literal_sample.append(
                                {
                                    "sheet": ws.title,
                                    "cell": cell.coordinate,
                                    "formula": formula,
                                }
                            )

                    if "[" in formula and "]" in formula:
                        external_reference_formula_count += 1
                        if len(external_sample) < 30:
                            external_sample.append(
                                {
                                    "sheet": ws.title,
                                    "cell": cell.coordinate,
                                    "formula": formula,
                                }
                            )

                    if not is_hidden_col and hidden_cols_set:
                        refs = set(COL_REF_RE.findall(upper_formula))
                        if refs.intersection(hidden_cols_set):
                            visible_formulas_referencing_hidden_cols += 1

                if isinstance(cached, str) and cached.startswith("#"):
                    sheet_error_cells += 1
                    error_cell_count += 1
                    error_by_type[cached] += 1
                    if len(error_sample) < 30:
                        error_sample.append(
                            {
                                "sheet": ws.title,
                                "cell": cell.coordinate,
                                "cached": cached,
                                "formula": (
                                    formula if isinstance(formula, str) else ""
                                ),
                            }
                        )

        sheet_stats.append(
            {
                "name": ws.title,
                "state": ws.sheet_state,
                "maxRow": ws.max_row,
                "maxCol": ws.max_column,
                "formulaCells": sheet_formula_cells,
                "visibleFormulaCells": sheet_visible_formula_cells,
                "errorCells": sheet_error_cells,
                "hiddenRows": sheet_hidden_rows,
                "hiddenCols": len(sheet_hidden_cols),
            }
        )

    summary = {
        "workbook": str(workbook_path),
        "sheetCount": len(wb_f.worksheets),
        "formulaCellCount": formula_cell_count,
        "visibleFormulaCells": visible_formula_cells,
        "hiddenRowCount": hidden_row_count,
        "hiddenColCount": hidden_col_count,
        "errorCellCount": error_cell_count,
        "errorByType": dict(error_by_type),
        "formulaRefLiteralCount": formula_ref_literal_count,
        "externalReferenceFormulaCount": external_reference_formula_count,
        "visibleFormulasReferencingHiddenCols": (
            visible_formulas_referencing_hidden_cols
        ),
    }

    risk = _build_risk(summary)
    recommendations = _build_recommendations(summary)
    parity_cells = _extract_parity_cells(wb_v)

    return {
        "summary": summary,
        "risk": risk,
        "recommendations": recommendations,
        "parityCells": parity_cells,
        "sheets": sheet_stats,
        "samples": {
            "formulaRefLiteral": formula_ref_literal_sample,
            "errors": error_sample,
            "externalReferences": external_sample,
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Audit CQT workbook structural robustness"
    )
    parser.add_argument(
        "--workbook",
        default=DEFAULT_WORKBOOK_PATH,
        help="Path to workbook (.xlsx/.xlsm)",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help="JSON output path. Ignored when --stdout is used.",
    )
    parser.add_argument(
        "--stdout",
        action="store_true",
        help="Print JSON to stdout",
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Exit non-zero when critical risk flags are found",
    )

    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    report = analyze_workbook(workbook_path)

    if args.stdout:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"CQT workbook audit generated at: {output_path}")
        print(
            f"Risk level: {report['risk']['level']} "
            f"(score={report['risk']['score']})"
        )

    if args.check and report["risk"]["checkFailed"]:
        check_error = (
            "CQT workbook audit check failed: " "critical structural flags detected."
        )
        print(
            check_error,
            file=sys.stderr,
        )
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
