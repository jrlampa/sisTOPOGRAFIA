from openpyxl import load_workbook

wb = load_workbook('CQTsimplificado_BECO DO MATA 7 - PARIDADE_FINAL.xlsx', data_only=False)
wr = wb['RAMAL']
wd = wb['DB']
wg = wb['GERAL']

print('RAMAL core cells')
for coord in [
    'AA14', 'AB14', 'AC14', 'AD14', 'AE14',
    'AA18', 'AA19', 'AA20', 'AA21', 'AA22', 'AA23', 'AA24', 'AA25', 'AA26', 'AA27', 'AA28', 'AA29', 'AA30',
    'AB18', 'AB19', 'AB20', 'AB21', 'AB22', 'AB23', 'AB24', 'AB25', 'AB26', 'AB27', 'AB28', 'AB29', 'AB30', 'AB35',
    'W16', 'X16', 'Y16'
]:
    print(coord, '=>', wr[coord].value)

print('\nRAMAL row 5 B:U (load profile factors)')
print([wr.cell(5, c).value for c in range(2, 22)])

print('\nRAMAL row 18 headers A:AE')
print([wr.cell(17, c).value for c in range(1, 32)])
print([wr.cell(18, c).value for c in range(1, 32)])

print('\nDB rows A30:B60 (TAB_CLAN_DMDI)')
for r in range(30, 61):
    print(r, wd[f'A{r}'].value, wd[f'B{r}'].value)

print('\nDB rows C30:F45 (AUX_CLAN start)')
for r in range(30, 46):
    print(r, [wd.cell(r, c).value for c in range(3, 7)])

print('\nGERAL key formulas around clandestino')
for coord in ['H2', 'I2', 'F9', 'G9', 'H9', 'I9', 'I10', 'I11', 'I12']:
    print(coord, '=>', wg[coord].value)

print('\nGERAL L2 and RAMAL AB31:AB40')
print('GERAL!L2 =>', wg['L2'].value)
for coord in ['AB31', 'AB32', 'AB33', 'AB34', 'AB35', 'AB36', 'AB37', 'AB38', 'AB39', 'AB40']:
    print(coord, '=>', wr[coord].value)
