from openpyxl import load_workbook

wb = load_workbook('CQTsimplificado_BECO DO MATA 7 - PARIDADE_FINAL.xlsx', data_only=False)

ws = wb['DB']
print('DB A30:F50')
for r in range(30, 51):
    print(r, [ws.cell(r, c).value for c in range(1, 7)])

print('\nDB A320:F330')
for r in range(320, 331):
    print(r, [ws.cell(r, c).value for c in range(1, 7)])

wr = wb['RAMAL']
print('\nRAMAL rows 17:30 cols W:AB (23..28)')
for r in range(17, 31):
    print(r, [wr.cell(r, c).value for c in range(23, 29)])

wg = wb['GERAL']
print('\nGERAL key formulas around clandestino')
for coord in ['H2', 'I2', 'F9', 'G9', 'H9', 'I9', 'I10', 'I11', 'I12']:
    print(coord, '=>', wg[coord].value)
