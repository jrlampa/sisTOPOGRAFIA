#!/usr/bin/env python3
"""
Extrator do catálogo de kits e materiais da Light.

Lê o arquivo CADASTRO DE KITS E MATERIAIS.xlsm e gera:
  - kits_catalog.json   (sheet KITS)
  - materiais_catalog.json (sheet KITS_ATUALIZADO)

Uso:
    python scripts/extract_kits_catalog.py [--input CAMINHO_XLSM] [--outdir DIR_SAIDA]

Defaults:
    --input:  C:/Users/jonat/OneDrive - IM3 Brasil/LIGHT/Arquivos para auxílio/CADASTRO DE KITS E MATERIAIS.xlsm
    --outdir: sisrua_unified/public/data/  (servido como asset estático)
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Erro: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

DEFAULT_INPUT = (
    r"C:\Users\jonat\OneDrive - IM3 Brasil\LIGHT\Arquivos para auxílio"
    r"\CADASTRO DE KITS E MATERIAIS.xlsm"
)
DEFAULT_OUTDIR = Path(__file__).resolve().parents[1] / "public" / "data"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _str(v) -> str:
    """Converte valor para string limpa, ou '' se nulo."""
    if v is None:
        return ""
    return str(v).strip()


def _int_or_str(v):
    """Retorna int se o valor já for int, senão str limpa."""
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v == int(v):
        return int(v)
    return _str(v)


# ---------------------------------------------------------------------------
# Extractors
# ---------------------------------------------------------------------------

def extract_kits(ws) -> list[dict]:
    """
    Sheet KITS — colunas: L.Tarefa, Alt, Oper, Descr
    Exclui a linha de cabeçalho (primeira linha não-nula).
    """
    kits: list[dict] = []
    header_skipped = False

    for row in ws.iter_rows(values_only=True):
        # Ignora linhas completamente vazias
        if not any(c for c in row if c is not None):
            continue

        # Primeira linha não-vazia é o cabeçalho
        if not header_skipped:
            header_skipped = True
            continue

        l_tarefa = row[0] if len(row) > 0 else None
        alt      = row[1] if len(row) > 1 else None
        oper     = row[2] if len(row) > 2 else None
        descr    = row[3] if len(row) > 3 else None

        # Ignora registros sem código de tarefa
        if l_tarefa is None:
            continue

        kits.append({
            "lTarefa": _int_or_str(l_tarefa),
            "alt":     int(alt) if isinstance(alt, (int, float)) and alt is not None else None,
            "oper":    int(oper) if isinstance(oper, (int, float)) and oper is not None else None,
            "descr":   _str(descr),
        })

    return kits


def extract_materiais(ws) -> list[dict]:
    """
    Sheet KITS_ATUALIZADO — colunas: RedStand, TxtBrv.LstTaref.
    """
    materiais: list[dict] = []
    header_skipped = False

    for row in ws.iter_rows(values_only=True):
        if not any(c for c in row if c is not None):
            continue

        if not header_skipped:
            header_skipped = True
            continue

        code  = row[0] if len(row) > 0 else None
        descr = row[1] if len(row) > 1 else None

        if code is None:
            continue

        materiais.append({
            "redStand": _int_or_str(code),
            "descricao": _str(descr),
        })

    return materiais


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Extrator catálogo KITS/MATERIAIS Light")
    parser.add_argument("--input",  default=DEFAULT_INPUT,  help="Caminho do .xlsm")
    parser.add_argument("--outdir", default=str(DEFAULT_OUTDIR), help="Diretório de saída")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"Erro: arquivo não encontrado: {src}")
        sys.exit(1)

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"Lendo: {src}")
    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)

    # --- KITS ---
    if "KITS" not in wb.sheetnames:
        print("Aviso: sheet 'KITS' não encontrada.")
        kits = []
    else:
        kits = extract_kits(wb["KITS"])
        print(f"  KITS:            {len(kits):>6} registros")

    # --- KITS_ATUALIZADO ---
    if "KITS_ATUALIZADO" not in wb.sheetnames:
        print("Aviso: sheet 'KITS_ATUALIZADO' não encontrada.")
        materiais = []
    else:
        materiais = extract_materiais(wb["KITS_ATUALIZADO"])
        print(f"  KITS_ATUALIZADO: {len(materiais):>6} registros")

    wb.close()

    # Grava kits_catalog.json
    kits_out = outdir / "kits_catalog.json"
    kits_out.write_text(
        json.dumps({"kits": kits}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Gerado: {kits_out}")

    # Grava materiais_catalog.json
    mat_out = outdir / "materiais_catalog.json"
    mat_out.write_text(
        json.dumps({"materiais": materiais}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Gerado: {mat_out}")

    print("Concluído.")


if __name__ == "__main__":
    main()
