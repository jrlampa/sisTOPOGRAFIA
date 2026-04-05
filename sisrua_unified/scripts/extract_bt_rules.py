from __future__ import annotations

from openpyxl import load_workbook

WORKBOOK_PATH = "CQTsimplificado_BECO DO MATA 7 - PARIDADE_FINAL.xlsx"

wb = load_workbook(WORKBOOK_PATH, data_only=False)

print("SHEETS")
for ws in wb.worksheets:
    print("-", ws.title)

print("\nDEFINED NAMES (matching BT rules)")
for name, defn in wb.defined_names.items():
    if any(k in name.upper() for k in ["AUX", "RAMAIS", "DMD", "MTTR", "GERAL", "CLAN", "TRAFO", "CQT"]):
        print(f"{name} => {defn}")

print("\nTABLES")
for ws in wb.worksheets:
    if ws.tables:
        print(f"[{ws.title}]")
        for tname, tbl in ws.tables.items():
            ref = getattr(tbl, "ref", tbl)
            print(f"  {tname}: {ref}")

# Inspect key sheets and anchors used by formulas.
for sheet_name in ["GERAL", "GERAL PROJ", "GERAL PROJ2", "AUX", "AUXILIAR", "DADOS"]:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f"\n[{sheet_name}] key cells")
    for coord in ["I2", "L2", "A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "A2", "B2", "C2", "D2", "E2", "F2", "G2", "H2"]:
        value = ws[coord].value
        if value is not None:
            print(f"  {coord}: {value}")

# Print formula pattern from GERAL PROJ row 9..18 relevant columns.
for sheet_name in ["GERAL PROJ", "GERAL PROJ2"]:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f"\n[{sheet_name}] formula pattern D:I row 9")
    for col in ["D", "E", "F", "G", "H", "I", "M"]:
        print(f"  {col}9: {ws[f'{col}9'].value}")

# Look for text clues in first 120 rows of all sheets.
keywords = ["CLANDEST", "CLT", "AUX_CLAN", "RAMAIS", "DMDI", "DEMANDA", "M2", "M²", "TRAFO"]
print("\nTEXTUAL CLUES")
for ws in wb.worksheets:
    hits = []
    for row in ws.iter_rows(min_row=1, max_row=min(120, ws.max_row), min_col=1, max_col=min(40, ws.max_column)):
        for c in row:
            v = c.value
            if isinstance(v, str):
                uv = v.upper()
                if any(k in uv for k in keywords):
                    hits.append((c.coordinate, v))
    if hits:
        print(f"[{ws.title}] {len(hits)} hits")
        for coord, v in hits[:30]:
            print(f"  {coord}: {v}")
