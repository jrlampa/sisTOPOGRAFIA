from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

WORKBOOK = "CQTsimplificado_BECO DO MATA 7 - PARIDADE_FINAL.xlsx"
OUT_FILE = Path("src/constants/clandestinoWorkbookRules.ts")

wb = load_workbook(WORKBOOK, data_only=True)
ws = wb["DB"]

area_to_kva: list[tuple[int, float]] = []
for row in range(31, 412):
    area = ws.cell(row, 1).value
    kva = ws.cell(row, 2).value
    if area is None or kva is None:
        continue
    area_to_kva.append((int(area), float(kva)))

client_to_diversif: list[tuple[int, float]] = []
for row in range(31, 331):
    clients = ws.cell(row, 3).value
    factor = ws.cell(row, 5).value
    if clients is None or factor is None:
        continue
    client_to_diversif.append((int(clients), float(factor)))

lines: list[str] = []
lines.append("// Auto-generated from workbook: CQTsimplificado_BECO DO MATA 7 - PARIDADE_FINAL.xlsx")
lines.append("// Source ranges: DB!A31:B411 (TAB_CLAN_DMDI), DB!C31:E330 (AUX_CLAN)")
lines.append("")
lines.append("export const CLANDESTINO_MIN_AREA_M2 = %d;" % min(a for a, _ in area_to_kva))
lines.append("export const CLANDESTINO_MAX_AREA_M2 = %d;" % max(a for a, _ in area_to_kva))
lines.append("export const CLANDESTINO_MIN_CLIENTS = %d;" % min(c for c, _ in client_to_diversif))
lines.append("export const CLANDESTINO_MAX_CLIENTS = %d;" % max(c for c, _ in client_to_diversif))
lines.append("")

lines.append("export const CLANDESTINO_AREA_TO_KVA: Record<number, number> = {")
for area, kva in area_to_kva:
    lines.append(f"  {area}: {kva},")
lines.append("};")
lines.append("")

lines.append("export const CLANDESTINO_CLIENT_TO_DIVERSIF_FACTOR: Record<number, number> = {")
for clients, factor in client_to_diversif:
    lines.append(f"  {clients}: {factor},")
lines.append("};")

OUT_FILE.write_text("\n".join(lines) + "\n", encoding="utf-8")
print(f"Generated {OUT_FILE} with {len(area_to_kva)} area rows and {len(client_to_diversif)} client rows")
