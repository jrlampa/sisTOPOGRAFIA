from __future__ import annotations

from openpyxl import load_workbook

WORKBOOK_PATH = "CQTsimplificado_BECO DO MATA 7 - PARIDADE_FINAL.xlsx"
KEYWORDS = [
    "clandest",
    "ramais",
    "geral",
    "demanda",
    "traf",
    "kwh",
    "m²",
    "m2",
    "cqt",
    "carga",
]


def contains_keyword(value: str) -> bool:
    text = value.lower()
    return any(keyword in text for keyword in KEYWORDS)


wb = load_workbook(WORKBOOK_PATH, data_only=False)
print("SHEETS:", wb.sheetnames)

for ws in wb.worksheets:
    hits: list[tuple[str, str]] = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and contains_keyword(value):
                hits.append((cell.coordinate, value))

    if hits:
        print(f"\n--- {ws.title} | hits: {len(hits)}")
        for coord, value in hits[:120]:
            print(f"{coord} => {value}")


print("\nFORMULAS WITH KEYWORDS IN TEXTUAL CONTEXT")
for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formula = value.lower()
                if any(key in formula for key in ["clandest", "ramais", "geral", "demanda", "kwh", "traf"]):
                    print(f"{ws.title}!{cell.coordinate} => {value}")
