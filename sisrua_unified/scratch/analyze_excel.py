import openpyxl
import os

file_path = r"C:\Users\jonat\OneDrive - IM3 Brasil\LIGHT\PROJETOS\CACUIA - RUA PEDREIRA - NOVA IGUAÇU\CQT - CACUIA - RUA PEDREIRA - NOVA IGUAÇU - ZNA1.xlsm"

def extract_logic(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        
        # Target sheets
        target_sheets = ['QDT Dutra 2.3 Lado Esquerdo', 'QDT Dutra 2.3 Lado Direito', 'Distrib. Cargas']
        
        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                print(f"Sheet {sheet_name} not found.")
                continue
                
            print(f"\n=== Analyzing Sheet: {sheet_name} ===")
            ws = wb[sheet_name]
            
            # Find header row
            header_row = 1
            for r in range(1, 15):
                if ws.cell(row=r, column=1).value and "PONTO" in str(ws.cell(row=r, column=1).value).upper():
                    header_row = r
                    break
            
            headers = [str(ws.cell(row=header_row, column=c).value).strip() if ws.cell(row=header_row, column=c).value else f"Col{c}" for c in range(1, 40)]
            print(f"Headers: {headers}")
            
            # Extract sample row (row 20 usually has data)
            data_row = header_row + 5
            print(f"\nSample Data (Row {data_row}):")
            for c, header in enumerate(headers, 1):
                cell = ws.cell(row=data_row, column=c)
                val = cell.value
                if val:
                    print(f"  {header}: {val}")
                    
        # Check constants
        if 'Tabela' in wb.sheetnames:
            print("\n=== Analyzing Constants (Tabela) ===")
            ws = wb['Tabela']
            for r in range(1, 30):
                row_vals = [str(ws.cell(row=r, column=c).value) for c in range(1, 10) if ws.cell(row=r, column=c).value]
                if row_vals:
                    print(f"Row {r}: {row_vals}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_logic(file_path)
