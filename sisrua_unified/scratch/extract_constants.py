import openpyxl
import os
import json

file_path = r"C:\Users\jonat\OneDrive - IM3 Brasil\LIGHT\PROJETOS\CACUIA - RUA PEDREIRA - NOVA IGUAÇU\CQT - CACUIA - RUA PEDREIRA - NOVA IGUAÇU - ZNA1.xlsm"

def extract_detailed_constants(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True) # Get values
        
        constants = {
            "cables": [],
            "transformers": []
        }
        
        if 'Tabela' in wb.sheetnames:
            ws = wb['Tabela']
            # Cable table starting row 7
            for r in range(7, 30):
                name = ws.cell(row=r, column=1).value
                if not name: continue
                r_ca = ws.cell(row=r, column=2).value # Rca (Z1, Z2)
                x_l = ws.cell(row=r, column=3).value  # jXL (Z1, Z2)
                r0 = ws.cell(row=r, column=4).value   # Rca (Z0)
                x0 = ws.cell(row=r, column=5).value   # jXL (Z0)
                
                constants["cables"].append({
                    "name": name,
                    "r": r_ca,
                    "x": x_l,
                    "r0": r0,
                    "x0": x0
                })
        
        # Look for Transformer data
        # Often in 'Base de Dados' or 'PP e CE'
        for sn in ['Base de Dados', 'PP e CE', 'Tabela']:
            if sn in wb.sheetnames:
                ws = wb[sn]
                # Search for KVA and Z
                for r in range(1, 50):
                    for c in range(1, 20):
                        val = str(ws.cell(row=r, column=c).value).upper()
                        if "KVA" in val and "Z" in val:
                             print(f"Trafo potential at {sn}!{r},{c}: {val}")

        print(json.dumps(constants, indent=2))

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_detailed_constants(file_path)
