import openpyxl
import os

file_path = r"C:\Users\jonat\OneDrive - IM3 Brasil\LIGHT\PROJETOS\CACUIA - RUA PEDREIRA - NOVA IGUAÇU\CQT - CACUIA - RUA PEDREIRA - NOVA IGUAÇU - ZNA1.xlsm"

def trace_deep_formulas(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb['QDT Dutra 2.3 Lado Esquerdo']
        
        # Scan for headers again, looking deeper
        headers = {}
        header_row = 1
        for r in range(10, 30):
            val = str(ws.cell(row=r, column=3).value).upper() # Col 3 often has the first "Ponto" or similar
            if "PONTO" in val or "DIST" in val:
                header_row = r
                break
        
        print(f"Header Row: {header_row}")
        for c in range(1, 40):
            val = ws.cell(row=header_row, column=c).value
            if val:
                headers[str(val).strip().upper()] = c
        
        print(f"Headers: {headers}")
        
        # Trace formulas at row 40 (likely has data)
        row_to_trace = 40
        print(f"\nFormula Trace at Row {row_to_trace}:")
        for h, col in headers.items():
            cell = ws.cell(row=row_to_trace, column=col)
            print(f"  {h} ({cell.coordinate}): {cell.value}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    trace_deep_formulas(file_path)
