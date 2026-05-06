import openpyxl
import os

file_path = r"C:\Users\jonat\OneDrive - IM3 Brasil\LIGHT\PROJETOS\CACUIA - RUA PEDREIRA - NOVA IGUAÇU\CQT - CACUIA - RUA PEDREIRA - NOVA IGUAÇU - ZNA1.xlsm"

def trace_formulas(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb['QDT Dutra 2.3 Lado Esquerdo']
        
        # Find headers
        headers = {}
        header_row = 1
        for r in range(1, 10):
            found = False
            for c in range(1, 30):
                val = str(ws.cell(row=r, column=c).value).upper()
                if "PONTO" in val or "CABO" in val or "IB" in val or "QUEDA" in val:
                    header_row = r
                    found = True
                    break
            if found: break
            
        for c in range(1, 40):
            val = ws.cell(row=header_row, column=c).value
            if val:
                headers[str(val).strip().upper()] = c
        
        print(f"Detected Headers: {list(headers.keys())}")
        
        data_row = header_row + 2 # Skip one maybe?
        if not ws.cell(row=data_row, column=headers.get('PONTO', 1)).value:
            data_row += 1
            
        print(f"\nTracing Formulas at Row {data_row}:")
        for h, col in headers.items():
            cell = ws.cell(row=data_row, column=col)
            print(f"  {h} ({cell.coordinate}): {cell.value}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    trace_formulas(file_path)
