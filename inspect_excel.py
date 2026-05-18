"""
Script para extrair dados de condutores, postes e métodos de cálculo do Excel.
"""

import sys
from openpyxl import load_workbook
from pathlib import Path
import json

# Caminho do arquivo
excel_path = Path(r"C:\Users\jonat\Downloads\CALC TRAÇÃO\GD - IMPACT CONSULTORIA LTDA - UFV ESTIMA__PARTE_3_POSTE_22.xlsm")

if not excel_path.exists():
    print(f"❌ Arquivo não encontrado: {excel_path}")
    sys.exit(1)

print(f"📂 Abrindo arquivo: {excel_path}")
print(f"📊 Tamanho: {excel_path.stat().st_size / 1024 / 1024:.2f} MB\n")

try:
    # Carregar workbook
    wb = load_workbook(excel_path, data_only=False)
    
    print("=" * 80)
    print("ABAS DISPONÍVEIS NO ARQUIVO")
    print("=" * 80)
    
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"\n{i}. {sheet_name}")
        print(f"   Linhas: {ws.max_row}, Colunas: {ws.max_column}")
        
        # Mostrar primeiras linhas
        print(f"   Primeiras linhas:")
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(8, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                if value is None:
                    value = ""
                row_data.append(str(value)[:20])
            print(f"     R{row_idx}: {' | '.join(row_data)}")
    
    print("\n" + "=" * 80)
    print("LENDO DADOS DETALHADOS")
    print("=" * 80)
    
    # Analisar cada aba
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n\n📋 PLANILHA: {sheet_name}")
        print("─" * 80)
        
        # Ler todos os dados
        data = []
        headers = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
            row_data = {}
            for col_idx, cell in enumerate(row, 1):
                col_letter = cell.column_letter
                value = cell.value
                
                # Guardar como header se primeira linha
                if row_idx == 1:
                    if headers is None:
                        headers = []
                    headers.append(value)
                
                row_data[f"C{col_idx}"] = value
            
            data.append(row_data)
            
            # Mostrar primeiras 15 linhas com detalle
            if row_idx <= 15:
                print(f"\nLinha {row_idx}:")
                for key, val in row_data.items():
                    if val is not None and str(val).strip():
                        print(f"  {key}: {val}")
        
        print(f"\nTotal de linhas em {sheet_name}: {len(data)}")
    
    # Análise específica por tipo de planilha
    print("\n" + "=" * 80)
    print("ANÁLISE ESPECÍFICA")
    print("=" * 80)
    
    # Procurar por palavras-chave
    keywords = {
        "condutor": ["al", "cu", "seção", "section", "mm2", "ohm", "resistance"],
        "poste": ["poste", "pole", "altura", "height", "effort", "esforço"],
        "tração": ["tração", "traction", "cálculo", "método", "formula", "tensão"],
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lower = sheet_name.lower()
        
        for category, terms in keywords.items():
            if any(term in sheet_lower for term in terms):
                print(f"\n✓ {sheet_name} => POTENCIAL FONTE DE DADOS: {category.upper()}")
    
    print("\n✅ Leitura completada. Próximo passo: análise focada dos dados.")
    
except Exception as e:
    print(f"❌ Erro ao abrir arquivo: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
